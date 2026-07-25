"""
Carga masiva de la nómina de trabajadores desde CSV o Excel.

Por qué existe: dar de alta un trabajador pide RUT, nombre y cargo. Un
contratista que parte con 80 personas abre ese diálogo 80 veces y escribe 240
campos. Es lo PRIMERO que hace el día 1, y es el momento en que decide si la
herramienta le ayuda o le agrega trabajo.

Tres decisiones que definen el comportamiento:

1. La importación es PARCIAL, no todo-o-nada. Con 80 filas siempre hay tres
   malas; rechazar el archivo completo obliga a repetir el trabajo bueno. Se
   cargan las válidas y se informa fila por fila qué pasó con las otras.

2. Un RUT repetido NO es un error, es "ya existía". Cargar dos veces el mismo
   archivo —corrigiendo tres filas— tiene que ser seguro. Sin esto, el segundo
   intento duplica a 77 personas.

3. El dígito verificador se valida. Es la razón de ser del validador: detecta el
   dígito cambiado de lugar, que es el error típico de transcribir a mano y que
   deja documentos colgando de alguien que no existe.
"""
import csv
import io
import unicodedata
from dataclasses import dataclass, field

from sqlalchemy.orm import Session

from app.core.exceptions import RutInvalido
from app.domain import rut_service
from app.models.trabajador import Trabajador

# Tope defensivo: una nómina real no pasa de unos cientos. Evita que un archivo
# equivocado —un export de 200k filas— tumbe el worker.
MAX_FILAS = 5000

COLUMNAS = ("rut", "nombre completo", "cargo")


@dataclass
class FilaError:
    fila: int
    rut: str
    nombre: str
    motivo: str


@dataclass
class ReporteImportacion:
    """Lo que se le muestra al usuario después de subir el archivo."""
    filas_leidas: int = 0
    cargados: int = 0
    ya_existian: int = 0
    errores: list[FilaError] = field(default_factory=list)

    @property
    def con_error(self) -> int:
        return len(self.errores)


def _sin_acentos(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", texto or "")
    return "".join(c for c in normalizado if not unicodedata.combining(c)).strip().lower()


def _decodificar(contenido: bytes) -> str:
    """
    Excel en Windows guarda CSV en cp1252, no en UTF-8. Si se asume UTF-8 a
    secas, un archivo con "Muñoz" revienta al decodificar y el usuario recibe un
    error de encoding que no significa nada para él.
    """
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return contenido.decode(encoding)
        except UnicodeDecodeError:
            continue
    return contenido.decode("utf-8", errors="replace")


def _filas_csv(contenido: bytes) -> list[list[str]]:
    texto = _decodificar(contenido)
    # Excel en español exporta con punto y coma. Detectarlo evita que toda la
    # fila caiga en una sola columna y el usuario vea "RUT vacío" en todo.
    muestra = texto[:4096]
    delimitador = ";" if muestra.count(";") > muestra.count(",") else ","
    return [f for f in csv.reader(io.StringIO(texto), delimiter=delimitador) if any(c.strip() for c in f)]


def _filas_xlsx(contenido: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "No se puede leer .xlsx en este servidor. Guarda la planilla como CSV y súbela de nuevo."
        )
    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro.active
    filas = []
    for fila in hoja.iter_rows(values_only=True):
        celdas = ["" if c is None else str(c).strip() for c in fila]
        if any(celdas):
            filas.append(celdas)
    return filas


def _indices_de_columnas(encabezado: list[str]) -> tuple[int, int, int]:
    """
    Ubica RUT / Nombre / Cargo por NOMBRE y no por posición: la gente reordena
    columnas en Excel, y leer por índice fijo cargaría los nombres en el RUT sin
    avisar. Devuelve -1 para cargo, que es opcional.
    """
    normalizado = [_sin_acentos(c) for c in encabezado]

    def buscar(*alias: str) -> int:
        for i, col in enumerate(normalizado):
            if col in alias:
                return i
        for i, col in enumerate(normalizado):
            if any(a in col for a in alias):
                return i
        return -1

    i_rut = buscar("rut")
    i_nombre = buscar("nombre completo", "nombre")
    i_cargo = buscar("cargo")
    if i_rut < 0 or i_nombre < 0:
        raise ValueError(
            "El archivo debe tener las columnas RUT y Nombre completo. "
            "Descarga la plantilla y úsala como base."
        )
    return i_rut, i_nombre, i_cargo


def plantilla_csv() -> bytes:
    """
    Plantilla de ejemplo. Lleva BOM porque sin él Excel abre el CSV en ANSI y
    parte los acentos; el usuario ve "Muñoz" y cree que la app está rota.

    Los RUT de ejemplo tienen dígito verificador VÁLIDO a propósito: si la
    plantilla trajera ejemplos inválidos, el primer intento de todo el mundo
    fallaría contra el propio archivo que les dimos.
    """
    salida = io.StringIO()
    escritor = csv.writer(salida, delimiter=";", lineterminator="\r\n")
    escritor.writerow(["RUT", "Nombre completo", "Cargo"])
    escritor.writerow(["17.555.666-4", "María Soto Vargas", "Prevencionista"])
    escritor.writerow(["19.444.555-5", "Pedro González Rojas", "Jefe de Obra"])
    escritor.writerow(["20.666.777-K", "Luis Muñoz Pérez", "Operador"])
    return "﻿".encode("utf-8") + salida.getvalue().encode("utf-8")


def importar_nomina(db: Session, empresa_id, contenido: bytes, nombre_archivo: str) -> ReporteImportacion:
    """
    Lee el archivo, valida cada fila y crea los trabajadores válidos.

    No hace commit por fila: se acumula y se confirma una vez al final, para que
    una nómina de 500 filas no sean 500 transacciones.
    """
    if nombre_archivo.lower().endswith((".xlsx", ".xlsm")):
        filas = _filas_xlsx(contenido)
    elif nombre_archivo.lower().endswith((".csv", ".txt")):
        filas = _filas_csv(contenido)
    else:
        raise ValueError("Formato no soportado. Sube un archivo .csv o .xlsx.")

    if not filas:
        raise ValueError("El archivo está vacío.")
    if len(filas) - 1 > MAX_FILAS:
        raise ValueError(f"El archivo tiene más de {MAX_FILAS} filas. Divídelo en partes.")

    i_rut, i_nombre, i_cargo = _indices_de_columnas(filas[0])
    reporte = ReporteImportacion()

    # Los RUT que ya tiene la empresa, en forma canónica. Una sola consulta: con
    # una por fila, 500 trabajadores son 500 viajes a la base.
    existentes = {
        rut_service.clave(r[0])
        for r in db.query(Trabajador.rut).filter(Trabajador.empresa_id == empresa_id).all()
    }
    vistos_en_archivo: set[str] = set()

    def celda(fila: list[str], i: int) -> str:
        return fila[i].strip() if 0 <= i < len(fila) else ""

    for numero, fila in enumerate(filas[1:], start=2):  # 1 es el encabezado
        reporte.filas_leidas += 1
        rut_crudo = celda(fila, i_rut)
        nombre = celda(fila, i_nombre)
        cargo = celda(fila, i_cargo) or None

        if not nombre:
            reporte.errores.append(FilaError(numero, rut_crudo, nombre, "Falta el nombre"))
            continue
        try:
            rut = rut_service.validar(rut_crudo)
        except RutInvalido as e:
            reporte.errores.append(FilaError(numero, rut_crudo, nombre, str(e)))
            continue

        canonico = rut_service.clave(rut)
        if canonico in existentes:
            reporte.ya_existian += 1
            continue
        if canonico in vistos_en_archivo:
            reporte.errores.append(
                FilaError(numero, rut_crudo, nombre, "RUT repetido dentro del mismo archivo")
            )
            continue

        vistos_en_archivo.add(canonico)
        db.add(Trabajador(
            empresa_id=empresa_id, rut=rut, nombre_completo=nombre, cargo=cargo, activo=True,
        ))
        reporte.cargados += 1

    db.commit()
    return reporte
